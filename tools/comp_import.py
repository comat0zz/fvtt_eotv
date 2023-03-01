#!/usr/bin/python3
## -*- coding: utf-8 -*-

import json
import os
import argparse 
import openpyxl
import string
import random

def id_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


archetypes = {
    "noble_female": "Аристократка",
    "cyborg_male": "Киборг",
    "companion": "Компаньонка",
    "courier_male": "Курьер",
    "mercenary_male": "Наёмник",
    "tactician_male": "Тактик",
    "daredevil": "Сорвиголова",
    "stranger": "Странник",
    "battlesuit": "Броненосец",
    "bountyhunter": "Охотник за головами",
    "supernova_female": "Сверхновая",
    "cenzor": "Цензор",
    "engineer": "Инженер",
    "doctor": "Доктор",
    "juggernaut_female": "Джаггернаут",
    "kinetick": "Кинетик",
    "digital_dao_male": "Дао Цифры",
    "emissary": "Эмиссар",
    "psychomant": "Психомант",
    "raelith": "Раэлит",
    "janissary": "Янычар",
    "scoundrel": "Авантюрист",
    "shadow": "Тень"
}

parser = argparse.ArgumentParser(description='Перегоняет xslx в JSON для подготовки компендиума')

parser.add_argument('--xlsx', 
    type=str, 
    help='Файл до эксельки', 
    dest = 'xlsx',
    required=True
)

args = parser.parse_args()

book = openpyxl.load_workbook(args.xlsx)
sheet = book.worksheets[0]

origins_tmp = {}

for row in range(1, 999):
    playbook = sheet.cell(row=row, column=1).value
    name = sheet.cell(row=row, column=2).value
    sign = sheet.cell(row=row, column=3).value
    desc = sheet.cell(row=row, column=4).value

    if name is None:
        break


    listItem = {
        "id": id_generator(8),
        "name": name,
        "sign": sign,
        "description": desc
    }

    if playbook in origins_tmp:
        origins_tmp[playbook]["data"]["list"].append(listItem)
    else:
        origins_tmp[playbook] = {
            "name": archetypes[playbook],
            "img": "",
            "folder": None,
            "type": "origins",
            "data": {
                "playbook": playbook,
                "list": [listItem]
            }
        }

origins = []
for origin in origins_tmp:
    origins.append(origins_tmp[origin])


cf1 = open('origins.json', 'w', encoding='utf-8')
cf1.writelines(json.dumps(origins, ensure_ascii=False, indent=4, sort_keys=True))
cf1.close()

'''
sheet = book.worksheets[1]

clusters = []
for row in range(2, 999):
    keyName = sheet.cell(row=row, column=1).value
    name = sheet.cell(row=row, column=2).value
    desc = sheet.cell(row=row, column=3).value

    if name is None:
        break

    clusters.append({
        "name": name,
        "img": "",
        "folder": None,
        "type": "clusters",
        "data": {
          "keyName": keyName,
          "description": desc
        }
    })

cf2 = open('clusters.json', 'w', encoding='utf-8')
cf2.writelines(json.dumps(clusters, ensure_ascii=False, indent=4, sort_keys=True))
cf2.close()

sheet = book.worksheets[2]
start_equips = []
for row in range(1, 999):

    playbook = sheet.cell(row=row, column=1).value
    name = sheet.cell(row=row, column=2).value
    desc = sheet.cell(row=row, column=3).value
    in_action = sheet.cell(row=row, column=4).value
    key = sheet.cell(row=row, column=5).value

    if name is None:
        break

    if in_action == 1:
        in_action = True
    else:
        in_action = False

    if key == "no":
        key = ""

    start_equips.append({
        "name": name,
        "img": "",
        "folder": None,
        "type": "equipments",
        "data": {
          "keyName": "",
          "playbook": playbook,
          "formula": key,
          "choose_one": in_action,
          "description": desc,
          "price": 0,
          "tags": []
        }
    })

cf3 = open('start_equips.json', 'w', encoding='utf-8')
cf3.writelines(json.dumps(start_equips, ensure_ascii=False, indent=4, sort_keys=True))
cf3.close()


sheet = book.worksheets[3]
tags = []
for row in range(2, 999):

    typeTag = sheet.cell(row=row, column=1).value
    keyName = sheet.cell(row=row, column=2).value
    name = sheet.cell(row=row, column=3).value
    description = sheet.cell(row=row, column=4).value

    if name is None:
        break

    tags.append({
        "name": name,
        "img": "",
        "folder": None,
        "type": "tags",
        "data": {
          "typeTag": typeTag,
          "keyName": keyName,
          "description": description
        }
    });


cf4 = open('tags.json', 'w', encoding='utf-8')
cf4.writelines(json.dumps(tags, ensure_ascii=False, indent=4, sort_keys=True))
cf4.close()
'''